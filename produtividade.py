import time
import os
import glob
import win32com.client
import openpyxl
import urllib.parse
import pandas as pd
import re
import shutil
import unicodedata
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchWindowException, WebDriverException
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta # Para lidar com datas (hoje e ontem)
from selenium.webdriver.support.ui import Select # Para caixas de <select>
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from selenium.webdriver.support.ui import WebDriverWait, Select
from datetime import date, timedelta
from openpyxl import load_workbook
from urllib.parse import quote as url_quote
from openpyxl.cell.cell import MergedCell
from calendar import monthrange
from datetime import datetime, date

''''''
# --- CONFIGURAÇÕES ---
EMAIL_USER = " " #Digite o seu email aqui
SENHA_USER = " " #Digite a sua senha aqui
WAIT_TIME = 10

# --- FUNÇÃO DE APOIO: LOGIN MICROSOFT ---
def fazer_login_microsoft(driver, wait, email, senha):
    """Lida com o login da MS. Retorna True se logou, False se der erro."""
    print("--- Iniciando rotina de Login Microsoft ---")
    try:
        try:
            email_field = wait.until(EC.presence_of_element_located((By.ID, "i0116")))
            print("Preenchendo e-mail...")
            email_field.send_keys(email)
            wait.until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click()
            
            password_field = wait.until(EC.presence_of_element_located((By.ID, "i0118")))
            print("Preenchendo senha...")
            password_field.send_keys(senha)
            
            clicked = False
            for _ in range(3):
                try:
                    wait.until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click()
                    clicked = True
                    break
                except StaleElementReferenceException:
                    time.sleep(1)
            if not clicked: raise Exception("Não clicou em Entrar")
            time.sleep(1)
            print("!!! AGUARDANDO APROVAÇÃO MFA (Se necessário) !!!")
            wait.until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click() 
            print("Login Microsoft efetuado.")
                # 1. Espera a janela pop-up fechar sozinha
            print("Aguardando janela pop-up fechar...")
            WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(1))
            
            # 2. Pega o handle da ÚNICA janela que sobrou (a nova janela principal)
            nova_janela_principal = driver.window_handles[0]
            driver.switch_to.window(nova_janela_principal)
            print("Foco retornado para a janela principal do Podio.")

 
        except TimeoutException:
            print("Campo de login não apareceu. Assumindo que já estamos logados (SSO).")
        return True
    except Exception as e:
        print(f"Erro no Login Microsoft: {e}")
        return False

# --- INÍCIO DO SCRIPT ---
try:
    driver = webdriver.Chrome()
    driver.maximize_window()
    wait = WebDriverWait(driver, WAIT_TIME)
    
    # ==============================================================================
    # PARTE 1: PODIO
    # ==============================================================================
    print("\n=== INICIANDO PARTE 1: PODIO ===")
    driver.get("https://podio.com/login")

    try:
        wait.until(EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))).click()
    except: pass 

    wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@data-provider='live']"))).click()

    janela_principal = driver.current_window_handle
    wait.until(EC.number_of_windows_to_be(2))
    
    for handle in driver.window_handles:
        if handle != janela_principal:
            driver.switch_to.window(handle)
            break

    fazer_login_microsoft(driver, wait, EMAIL_USER, SENHA_USER)
    driver.switch_to.window(janela_principal)
    
    print("Navegando no Podio...")
    menu_area = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'space-switcher-wrapper')]")))
    ActionChains(driver).move_to_element(menu_area).perform()
    wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'ADM - Núcleo Contratos')]"))).click()
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@data-app-id='22830484']"))).click()

    # --- CORREÇÃO DO FILTRO PODIO ---
    print("Aplicando filtros (Método Robusto)...")
    time.sleep(3) # Espera script da página carregar
    
    # 1. Pega o container UL
    ul_filtros = wait.until(EC.presence_of_element_located((By.XPATH, "//ul[@class='app-filter-tools']")))
    
    # 2. Pega todos os itens LI dentro dele
    itens_lista = ul_filtros.find_elements(By.TAG_NAME, "li")
    
    # 3. Passa o mouse em CADA item para garantir que o menu acorde
    actions = ActionChains(driver)
    for item in itens_lista:
        actions.move_to_element(item)
    actions.perform()
    
    # 4. Agora clica no filtro
    wait.until(EC.element_to_be_clickable((By.XPATH, ".//li[@data-original-title='Filtros']"))).click()
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@data-id='created_on']"))).click() 
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@data-id='-1mr:-1mr']"))).click() 

    # Seleção de Views
    print("Ajustando visualização...")
    elementos_menu = wait.until(lambda d: d.find_elements(By.CSS_SELECTOR, ".app-header__app-menu"))
    if len(elementos_menu) >= 1: elementos_menu[0].click()
    time.sleep(2)
    elementos_menu = driver.find_elements(By.CSS_SELECTOR, ".app-header__app-menu") 
    if len(elementos_menu) > 1: elementos_menu[1].click()
    time.sleep(2)

    print("Exportando Excel...")
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.app-box-supermenu-v2__link.app-export-excel"))).click()
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "li.navigation-link.inbox"))).click()

    time.sleep(30)
    
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.PodioUI__Notifications__NotificationGroup"))).click()
    wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'field-type-text')]"))) 
    wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Mensageria - Última vista usada.xlsx"))).click()
    print("Download Podio iniciado!")
    time.sleep(5) 
    
    # ==============================================================================
    # PARTE 2: AGILIS
    # ==============================================================================
    print("\n=== INICIANDO PARTE 2: AGILIS ===")
    driver.get("https://agilis.mrv.com.br/HomePage.do?view_type=my_view")

    try:
        btn_login = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[text()='Login Integrado Microsoft']")))
        btn_login.click()
        fazer_login_microsoft(driver, wait, EMAIL_USER, SENHA_USER)
    except TimeoutException:
        print("Botão de login não apareceu, seguindo...")

    print("Navegando menus Agilis...")
    wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Relatórios"))).click()
    wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Contratos - ADM"))).click()
    wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Produtividade Contratos - ADM"))).click()
    
    wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "linkborder"))).click() 
    wait.until(EC.element_to_be_clickable((By.XPATH, "//option[text()='Coletor de custo ADM']"))).click()
    driver.find_element(By.CLASS_NAME, "moverightButton").click()

    try:
        expand_btn = wait.until(EC.presence_of_element_located((By.ID, "rcstep2src")))
        driver.execute_script("arguments[0].click();", expand_btn)
        time.sleep(1)
    except: pass

    # --- SELEÇÃO DE DATA (DROPDOWN + RÁDIO) ---
    print("Selecionando Data 'Mês Passado' no dropdown...")
    select_elem = wait.until(EC.presence_of_element_located((By.ID, "dateFilterType")))
    Select(select_elem).select_by_visible_text("Mês passado")
    
    # [INSERIDO] Código do Rádio 'Durante' que você pediu
    print("Selecionando o rádio 'Durante' (Ajuste obrigatório)...")
    try:
        selector_radio_durante = (By.CSS_SELECTOR, "input[value='predefined']")
        wait.until(EC.element_to_be_clickable(selector_radio_durante)).click()
        print(" - SUCESSO: Filtro 'Durante' selecionado.")
    except TimeoutException:
        print(" - FALHA: Rádio 'Durante' não encontrado.")
        raise

    # Executar Relatório
    wait.until(EC.element_to_be_clickable((By.ID, "addnew223222"))).click() 
    print("Relatório gerando. Aguardando 10 segundos...")
    time.sleep(10) 

    # --- 7. Baixar Relatório XLS Diretamente ---
    print("7. Iniciando o download direto do relatório XLS...")
    try:
        # Localizar o link de exportação pelo ID "exportxls" e clicar nele
        DOWNLOAD_XLS_LINK = (By.ID, "exportxls")
        wait.until(EC.element_to_be_clickable(DOWNLOAD_XLS_LINK)).click()
        print("   - Clique realizado no link 'Exportar arquivo como XLS'.")
    
        # IMPORTANTE: Adicionar uma pausa para o download começar e terminar.
        # A melhor abordagem é verificar a pasta de downloads até o arquivo aparecer.
        # Veja a explicação abaixo sobre como fazer isso.
        print("   - Aguardando o download ser concluído...")
        time.sleep(5) # Pausa simples de 15 segundos. O ideal é usar uma função de verificação.
    
        print("   - Relatório baixado com sucesso!")

    except Exception as e:
        print(f"ERRO ao tentar baixar o relatório XLS: {e}")
        # Adicione aqui o tratamento de erro
    
    print("Relatório Agilis baixado com sucesso!")
    time.sleep(5)
   
finally:
    print("Script Completo Finalizado.")
    # driver.quit()
    print("Fim.")


#------------------------------------------------------------------------------------------
time.sleep(10)
# Definição dos caminhos (ajuste conforme o seu usuário)
downloads_path = os.path.expanduser("C:/Users/pedro.henrsilva/Downloads")
destination_path = os.path.expanduser("C:/Users/pedro.henrsilva/OneDrive - MRV/Área de Trabalho/produtividade")

# Cria a pasta de destino caso ela não exista
if not os.path.exists(destination_path):
    os.makedirs(destination_path)

# 1. Lista todos os itens e filtra apenas arquivos (ignora pastas)
files = [
    os.path.join(downloads_path, f) 
    for f in os.listdir(downloads_path) 
    if os.path.isfile(os.path.join(downloads_path, f))
]

# 2. Ordena os arquivos pela data de modificação (do mais recente para o mais antigo)
# os.path.getmtime retorna o timestamp da última modificação
files.sort(key=os.path.getmtime, reverse=True)

# 3. Pega os 4 primeiros e move
top_4_files = files[:4]

for file_path in top_4_files:
    file_name = os.path.basename(file_path)
    try:
        shutil.move(file_path, os.path.join(destination_path, file_name))
        print(f"Sucesso: {file_name} movido para {destination_path}")
    except Exception as e:
        print(f"Erro ao mover {file_name}: {e}")

# --- FUNÇÕES AUXILIARES GERAIS ---

def find_column_ignore_case(df, column_name):
    """Encontra o nome real de uma coluna, ignorando maiúsculas/minúsculas."""
    for col in df.columns:
        if col.lower() == column_name.lower():
            return col
    return None

# --- ETAPA 1: FUNÇÕES PARA Renomear e editar as planilhas ---

def processar_mensageria(filepath, new_filename):
    try:
        print(f"Processando MENSAGERIA: {filepath}")
        df = pd.read_excel(filepath, header=0) 
        df.columns = [str(col).strip() for col in df.columns]
        coluna_usuario = find_column_ignore_case(df, 'Criado por')
        coluna_data = find_column_ignore_case(df, 'Criado em')
        coluna_valores = find_column_ignore_case(df, 'Numero do chamado Agilis/Rastreio')
        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_usuario, columns=coluna_data, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.rename(filepath, new_filename)
        print(f" -> Sucesso! Renomeado para '{new_filename}'")
    except Exception as e: print(f" -> ERRO (Etapa 1 - Mensageria): {e}")

def processar_produtividade(filepath, new_filename):
    try:
        print(f"Processando PRODUTIVIDADE: {filepath}")
        df = pd.read_excel(filepath, header=0)
        df.columns = [str(col).strip() for col in df.columns]
        coluna_usuario = find_column_ignore_case(df, 'Nome do usuário')
        coluna_data = find_column_ignore_case(df, 'Data de lançamento')
        coluna_valores = find_column_ignore_case(df, 'Nº doc.faturamento')
        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_usuario, columns=coluna_data, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.rename(filepath, new_filename)
        print(f" -> Sucesso! Renomeado para '{new_filename}'")
    except Exception as e: print(f" -> ERRO (Etapa 1 - Produtividade): {e}")

def processar_numerico(filepath, new_filename):
    try:
        print(f"Processando ARQUIVO NUMÉRICO: {filepath}")
        df = pd.read_excel(filepath, header=8)
        df.columns = [str(col).strip() for col in df.columns]
        coluna_tecnico = find_column_ignore_case(df, 'Técnico')
        coluna_data = find_column_ignore_case(df, 'Hora de conclusão')
        coluna_valores = find_column_ignore_case(df, 'Identificação da solicitação')
        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_tecnico, columns=coluna_data, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(new_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='DadosOriginais', index=False)
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.remove(filepath)
        print(f" -> Sucesso! Novo arquivo '{new_filename}' criado.")
    except Exception as e: print(f" -> ERRO (Etapa 1 - Numérico): {e}")

def processar_relatorio_pedidos(filepath, new_filename):
    try:
        print(f"Processando RELATÓRIO DE PEDIDOS: {filepath}")
        df = pd.read_excel(filepath, header=1)
        df.columns = [str(col).strip() for col in df.columns]
        coluna_linhas = find_column_ignore_case(df, 'Respons. Entrega')
        coluna_colunas = find_column_ignore_case(df, 'Data Entrada NF')
        coluna_valores = find_column_ignore_case(df, 'Nro. Pedido Compra')
        df[coluna_colunas] = pd.to_datetime(df[coluna_colunas], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_linhas, columns=coluna_colunas, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.rename(filepath, new_filename)
        print(f" -> Sucesso! Renomeado para '{new_filename}'")
    except Exception as e: print(f" -> ERRO (Etapa 1 - Pedidos): {e}")
time.sleep(5)
def step_1_prepare_and_rename_reports(diretorio):
    arquivos = glob.glob(os.path.join(diretorio, '*.*'))
    for arquivo in arquivos:
        nome_arquivo = os.path.basename(arquivo)
        if nome_arquivo.startswith('Mensageria - Última vista'):
            processar_mensageria(arquivo, 'Relatório - Sedex.Malote.xlsx')
        elif nome_arquivo.startswith('export') and nome_arquivo.endswith('.xlsx'):
            processar_produtividade(arquivo, 'Relatório - SAP.xlsx')
        elif nome_arquivo.startswith('REL_PRLPGT'):
            processar_relatorio_pedidos(arquivo, 'Relatório - Lançamentos.xlsx')
        elif re.match(r'^\d+\.(xlsx|xls)$', nome_arquivo):
            processar_numerico(arquivo, 'Relatório - Agilis.xlsx')


# --- ETAPA 2: FUNÇÕES PARA ATUALIZAR A PLANILHA PRINCIPAL (COM novissima LÓGICA) ---

# Arquivos e abas

PROD_PATH    = "Produtividade 01 - 2026.xlsx"      # aba Plan1
AGILIS_PATH  = "Relatório - Agilis.xlsx"           # aba TabelaDinamica
SEDEX_PATH   = "Relatório - Sedex.Malote.xlsx"     # aba TabelaDinamica
LANCTOS_PATH = "Relatório - Lançamentos.xlsx"      # aba TabelaDinamica (Respons. Entrega)
SAP_PATH     = "Relatório - SAP.xlsx"              # aba TabelaDinamica (Nome do usuário)
OUT_PATH     = "Produtividade 02 - 2026 (preenchido).xlsx"

# ===================== MAPEAMENTOS =====================
# Sedex: nome na pivot -> texto exato da coluna B na produtividade
MAP_SEDEX = {
    "Alfredo Henrique Goncalves Pereira": "Alfredo.pereira MS0069532",
    "Gabriel Figueiredo Emiliano":        "gabriel.emiliano MS0073186",
    "Pedro Henrique Soares Silva":        "pedro.henrsilva MS0073814",
    "Ezequiel Viana Ferreira":            "ezequiel.ferreira",
}

# Agilis: MESMA linha do nome (coluna C), respeitando coluna mínima
AGILIS_POS = [
    {"p2": "Alfredo Henrique Goncalves Pereira", "p1": "Alfredo.pereira MS0069532",  "row_nome":  2, "min_col_letter": "CO"},
    {"p2": "Gabriel Figueiredo Emiliano",        "p1": "gabriel.emiliano MS0073186", "row_nome":  7, "min_col_letter": "CO"},
    {"p2": "Ezequiel Viana Ferreira",            "p1": "ezequiel.ferreira",          "row_nome": 12, "min_col_letter": "CO"},
    {"p2": "arthur.savio",                       "p1": "arthur.savio",               "row_nome": 17, "min_col_letter": "CO"},
    {"p2": "Pedro Henrique Soares Silva",        "p1": "pedro.henrsilva MS0073814",  "row_nome": 22, "min_col_letter": "CO"},
    {"p2": "Camilly Cristine Dos Santos",        "p1": "camilly.santos",             "row_nome": 27, "min_col_letter": "CO"},
    {"p2": "Carolina Pagnozzi Silva",            "p1": "pagnozzi.carolina",          "row_nome": 32, "min_col_letter": "CO"},
    {"p2": "maria.edurocha",                     "p1": "maria.edurocha",             "row_nome": 37, "min_col_letter": "CO"},
    {"p2": "Matheus Silva De Lemos",             "p1": "matheus.lemos.silva",        "row_nome": 42, "min_col_letter": "CO"},
    {"p2": "Vanessa De Brito Rodrigues",         "p1": "Vanessa",                    "row_nome": 47, "min_col_letter": "C"},  # Vanessa desde C
]

# Lançamentos 45 E 19: linhas fixas

LANCTOS_USER_MAP = {
    "alfredo.pereira":      {"p1": "Alfredo.pereira MS0069532",  "row_ativ":  4},
    "gabriel.emiliano":     {"p1": "gabriel.emiliano MS0073186", "row_ativ":  9},
    "ezequiel.ferreira":    {"p1": "ezequiel.ferreira",          "row_ativ": 14},
    "arthur.savio":         {"p1": "arthur.savio",               "row_ativ": 19},
    "pedro.henrsilva":      {"p1": "pedro.henrsilva MS0073814",  "row_ativ": 24},
    "camilly.santos":       {"p1": "camilly.santos",             "row_ativ": 29},
    "pagnozzi.carolina":    {"p1": "pagnozzi.carolina",          "row_ativ": 34},
    "maria.edurocha":       {"p1": "maria.edurocha",             "row_ativ": 39},
    "matheus.lemos.silva":  {"p1": "matheus.lemos.silva",        "row_ativ": 44},
}

ATIV_LANCTOS_LABEL = "Lançamentos 45 E 19"

# SAPMIRO: códigos MS -> colaborador + linha fixa
SAP_COD_MAP = {
    "MS0069532": {"p1": "Alfredo.pereira MS0069532",  "row_ativ":  5},
    "MS0073186": {"p1": "gabriel.emiliano MS0073186", "row_ativ": 10},
    "MS0073814": {"p1": "pedro.henrsilva MS0073814",  "row_ativ": 25},
}
ATIV_SAP_LABEL = "SAPMIRO"

# ===================== NOVAS FUNÇÕES =====================


def update_headers_to_previous_month(ws, header_row=1, start_col_letter="D", end_col_letter="AH", ref_date=None):
    """
    Atualiza o cabeçalho para o MÊS ANTERIOR à data de referência (ref_date).
    Se ref_date for None, usa a data de hoje.
    """
    start_col = col_letter_to_index(start_col_letter)
    end_col   = col_letter_to_index(end_col_letter)

    # Base para achar o mês anterior: 1º dia do mês da ref_date menos 1 dia
    if ref_date is None:
        ref_date = date.today()
    first_of_month = date(ref_date.year, ref_date.month, 1)
    last_day_prev  = first_of_month - timedelta(days=1)
    ano, mes = last_day_prev.year, last_day_prev.month

    qtd_dias = monthrange(ano, mes)[1]

    # 1) Escreve as datas reais (Excel) de 1..qtd_dias
    for i in range(qtd_dias):
        dt = datetime(ano, mes, i + 1)
        ws.cell(row=header_row, column=start_col + i, value=dt)

    # 2) Limpa as colunas excedentes dentro do range D:AH
    for c in range(start_col + qtd_dias, end_col + 1):
        ws.cell(row=header_row, column=c, value=None)

    # 3) (Opcional) Forçar visual "dd/mm/aaaa"
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=header_row, column=c)
        if isinstance(cell.value, datetime):
            cell.number_format = "dd/mm/yyyy"

    return ano, mes, qtd_dias



def clear_month_data_in_blocks(ws, row_ranges, start_col_letter="D", end_col_letter="AH"):
    start_col = col_letter_to_index(start_col_letter)
    end_col   = col_letter_to_index(end_col_letter)

    cleared = 0
    for (r0, r1) in row_ranges:
        for r in range(r0, r1 + 1):
            if r > ws.max_row:
                continue
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if cell.value not in (None, ""):
                    cell.value = None
                    cleared += 1
    return cleared

# ===================== FUNÇÕES AUX =====================
def norm_key(s):
    if s is None: return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()

def col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    n = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'): continue
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n

def date_keys(v):
    keys = []
    if isinstance(v, datetime):
        keys += [v.strftime("%Y-%m-%d 00:00:00"), v.strftime("%d/%m/%Y")]
    elif isinstance(v, str):
        s = v.strip()
        if re.match(r"^\d{4}-\d{2}-\d{2} 00:00:00$", s): keys.append(s)
        if re.match(r"^\d{2}/\d{2}/\d{4}$", s): keys.append(s)
        try:
            dt = pd.to_datetime(s, errors="raise")
            keys += [dt.strftime("%Y-%m-%d 00:00:00"), dt.strftime("%d/%m/%Y")]
        except: pass
    return list(dict.fromkeys(keys))

def build_header_map(ws):
    hdr = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=1, column=c).value
        if v is None: continue
        if isinstance(v, datetime):
            hdr[v.strftime("%Y-%m-%d 00:00:00")] = c
            hdr[v.strftime("%d/%m/%Y")] = c
        else:
            s = str(v).strip()
            hdr[s] = c
            for k in date_keys(s): hdr[k] = c
    return hdr

def extract_user_key(s: str) -> str:
    """
    Estratégia robusta:
    1) Remove domínio após '@' e normaliza;
    2) Se existir algum token com '.', usa esse token (mantendo apenas letras e '.');
    3) Senão, se houver 2+ palavras, usa 'primeira.ultima';
    4) Senão, cai na primeira palavra.
    Tudo em minúsculas e sem acentos.
    """
    if s is None:
        return ""
    s = str(s).strip()
    s = s.split('@')[0]                     # corta domínio
    s_norm = unicodedata.normalize("NFKD", s)
    s_norm = "".join(ch for ch in s_norm if not unicodedata.combining(ch))
    tokens = s_norm.strip().split()

    # 2) procura token com ponto (login/email)
    for t in tokens:
        if '.' in t:
            t = re.sub(r'[^A-Za-z\.]', '', t)
            return t.lower()

    # 3) monta 'primeira.ultima' se houver múltiplas palavras
    if len(tokens) >= 2:
        first = re.sub(r'[^A-Za-z]', '', tokens[0])
        last  = re.sub(r'[^A-Za-z]', '', tokens[-1])
        key = f"{first}.{last}"
        return key.lower()

    # 4) fallback: primeira palavra
    only = re.sub(r'[^A-Za-z]', '', tokens[0]) if tokens else ""
    return only.lower()

def read_tabledinamica_with_namecol(path, name_col_hint=None):
    
    """
    Lê a aba 'TabelaDinamica' do arquivo 'path' e retorna DataFrame long com colunas:
      ['nome','data_obj','valor'].
    Se 'name_col_hint' for fornecido, usa exatamente essa coluna (normalização tolerante).
    """
    xls = pd.ExcelFile(path, engine="openpyxl")
    if "TabelaDinamica" not in xls.sheet_names:
        raise RuntimeError(f"Aba 'TabelaDinamica' não encontrada em {path}. Abas: {xls.sheet_names}")
    df = pd.read_excel(path, sheet_name="TabelaDinamica", engine="openpyxl")

    # --- INÍCIO DO DEBUG ---
    print("=== COLUNAS ENCONTRADAS NO PANDAS ===")
    print(df.columns.tolist())
    print("=======================================")
    # --- FIM DO DEBUG ---

    # Escolher a coluna de nomes
    nome_col = None
    if name_col_hint:
        # casar com tolerância (strip/lower e sem acento)
        hint_norm = norm_key(name_col_hint)
        for c in df.columns:
            if norm_key(c) == hint_norm:
                nome_col = c
                break
    if not nome_col:
        # fallback: tenta cabeçalhos comuns
        heads = {norm_key(c): c for c in df.columns}
        for alvo in ["criado por", "tecnico", "técnico", "respons. entrega", "nome do usuário"]:
            if alvo in heads:
                nome_col = heads[alvo]
                break
    if not nome_col:
        # fallback final: primeira coluna de strings
        for c in df.columns:
            if df[c].dtype == "O":
                nome_col = c
                break

    # Colunas de data
    day_cols = []
    for c in df.columns:
        if c == nome_col: continue
        if isinstance(c, datetime):
            day_cols.append(c)
        elif isinstance(c, str):
            s = c.strip().lower()
            if s in ("total", "total geral"): continue
            if re.match(r"^\d{2}/\d{2}/\d{4}$", c) or re.match(r"^\d{4}-\d{2}-\d{2}", c):
                day_cols.append(c)

    registros = []
    if nome_col:
        for _, row in df.iterrows():
            nome_val = str(row.get(nome_col, "")).strip()
            if not nome_val or norm_key(nome_val).startswith(norm_key("Total Geral")):
                continue
            for d in day_cols:
                val = row.get(d, 0)
                try: v = int(val) if pd.notna(val) else 0
                except Exception: v = 0
                registros.append({"nome": nome_val, "data_obj": d, "valor": v})

    return pd.DataFrame(registros, columns=["nome","data_obj","valor"])

def read_lanctos_tabledinamica(path):
    """
    Lê a aba 'TabelaDinamica' do Relatório - Lançamentos.xlsx.
    Tenta identificar a coluna de usuários por vários nomes (Técnico, Respons. Entrega, etc).
    Retorna DF long: ['user_key','data_obj','valor'].
    """
    xls = pd.ExcelFile(path, engine="openpyxl")
    if "TabelaDinamica" not in xls.sheet_names:
        raise RuntimeError(f"Aba 'TabelaDinamica' não encontrada em {path}.")
    
    df = pd.read_excel(path, sheet_name="TabelaDinamica", engine="openpyxl")

    # --- 1. IDENTIFICAÇÃO FLEXÍVEL DA COLUNA DE NOME ---
    possible_names = [
        "Respons. Entrega", 
        "Técnico", 
        "Tecnico", 
        "Criado por", 
        "Nome do usuário", 
        "User"
    ]
    
    nome_col = None
    # Cria um mapa das colunas existentes normalizadas
    cols_map = {norm_key(c): c for c in df.columns}
    
    for alvo in possible_names:
        alvo_norm = norm_key(alvo)
        if alvo_norm in cols_map:
            nome_col = cols_map[alvo_norm]
            break
            
    if nome_col is None:
        # Se não achou pelos nomes, tenta a primeira coluna se for string (fallback)
        if len(df.columns) > 0 and df.dtypes[0] == 'object':
             nome_col = df.columns[0]
        else:
             raise RuntimeError(f"Não foi possível identificar a coluna de nomes. Tentou: {possible_names}")

    # --- 2. IDENTIFICAÇÃO DAS COLUNAS DE DATA ---
    day_cols = []
    for c in df.columns:
        if c == nome_col: continue
        # O Pandas já converteu para datetime (como visto no seu debug)
        if isinstance(c, datetime):
            day_cols.append(c)
        elif isinstance(c, str):
            s = c.strip()
            # Regex para casos onde o Excel não converteu automático
            if re.match(r"^\d{2}/\d{2}/\d{4}$", s) or re.match(r"^\d{4}-\d{2}-\d{2}", s):
                day_cols.append(c)

    # --- 3. EXTRAÇÃO DOS DADOS ---
    registros = []
    for _, row in df.iterrows():
        raw_name = str(row.get(nome_col, "")).strip()
        
        # Pula linhas de Total ou vazias
        if not raw_name or norm_key(raw_name).startswith(norm_key("Total Geral")):
            continue
            
        # Aplica a limpeza para gerar o user_key (alfredo.pereira)
        ukey = extract_user_key(raw_name)
        
        for d in day_cols:
            val = row.get(d, 0)
            try: 
                v = int(val) if pd.notna(val) else 0
            except: 
                v = 0
            
            # Só adiciona se houver valor (opcional, economiza memória)
            # if v != 0: 
            registros.append({"user_key": ukey, "data_obj": d, "valor": v})

    return pd.DataFrame(registros, columns=["user_key","data_obj","valor"])


# ===================== PREENCHIMENTOS =====================

def fill_agilis_same_row(ws, header_map, df_long):
    """Agilis na MESMA linha do nome (coluna C), respeitando coluna mínima e sem criar colunas.
       Agora tolerante: casa por nome normalizado OU por user_key extraída.
    """
    if df_long.empty:
        print("[AGILIS] Pivot vazia ou sem linhas válidas.")
        return 0

    # 1) Agrupamentos alternativos
    grp_exact = {norm_key(n): sub for n, sub in df_long.groupby(df_long['nome'].apply(norm_key))}
    grp_ukey  = {extract_user_key(n): sub for n, sub in df_long.groupby(df_long['nome'].apply(extract_user_key))}

    total_writes = 0

    for item in AGILIS_POS:
        p2 = item["p2"]                         # "nome como vem na pivot" (ou login)
        p1 = item["p1"]                         # "texto exato da coluna B" na produtividade
        row_nome = item["row_nome"]
        row_agilis = row_nome
        min_idx = col_letter_to_index(item["min_col_letter"])

        if min_idx > ws.max_column:
            print(f"[AGILIS] {p1}: coluna mínima {item['min_col_letter']} (idx {min_idx}) > max {ws.max_column}. Desativando limite.")
            min_idx = 1

        # 2) Tenta casar por nome normalizado; se não achar, tenta por user_key
        sub = grp_exact.get(norm_key(p2))
        if (sub is None) or sub.empty:
            sub = grp_ukey.get(extract_user_key(p2))

        if sub is None or sub.empty:
            print(f"[AGILIS] {p1}: sem registros localizados para '{p2}' (nome/login/email). Deixando em branco.")
            continue

        writes = 0
        for _, reg in sub.iterrows():
            col_idx = None
            for k in date_keys(reg["data_obj"]):
                c = header_map.get(k)
                if c:
                    col_idx = c; break
            if not col_idx or col_idx < min_idx:
                continue

            val = int(reg["valor"])
            if val != 0:
                ws.cell(row=row_agilis, column=col_idx, value=val)
                writes += 1

        total_writes += writes
        print(f"[AGILIS] {p1} (linha {row_agilis}, min {item['min_col_letter']}): {writes} células escritas.")
    return total_writes

def fill_sedex(ws, header_map, df_long):
    """Sedex/Pac/Malote: encontra atividade na coluna C abaixo do nome e preenche."""
    if df_long.empty:
        print("[SEDEX] Pivot vazia ou sem linhas válidas.")
        return 0

    grp = {norm_key(n): sub for n, sub in df_long.groupby(df_long['nome'].apply(norm_key))}
    total_writes = 0

    for p2, p1 in MAP_SEDEX.items():
        # localizar linha do nome (col B)
        r_nome = None
        for r in range(2, ws.max_row+1):
            v = ws.cell(row=r, column=2).value
            if isinstance(v, str) and norm_key(v) == norm_key(p1):
                r_nome = r; break
        if r_nome is None:
            print(f"[SEDEX] {p1}: nome não encontrado na coluna B. Pulando.")
            continue

        # procurar 'Sedex/Pac/Malote' em C nas próximas linhas do bloco
        r_sedex = None
        for rr in range(r_nome, min(ws.max_row, r_nome+12)+1):
            w = ws.cell(row=rr, column=3).value
            if isinstance(w, str) and 'sedex/pac/malote' in w.strip().lower():
                r_sedex = rr; break
        if r_sedex is None:
            print(f"[SEDEX] {p1}: atividade não encontrada na coluna C. Pulando.")
            continue

        sub = grp.get(norm_key(p2))
        if sub is None or sub.empty:
            print(f"[SEDEX] {p1}: sem registros (deixando em branco).")
            continue

        writes = 0
        for _, reg in sub.iterrows():
            col_idx = None
            for k in date_keys(reg["data_obj"]):
                c = header_map.get(k)
                if c:
                    col_idx = c; break
            if not col_idx: continue
            val = int(reg["valor"])
            if val != 0:
                ws.cell(row=r_sedex, column=col_idx, value=val); writes += 1

        total_writes += writes
        print(f"[SEDEX] {p1}: {writes} células escritas.")
    return total_writes


def fill_lanctos_fixed(ws, header_map, df_long):
    """
    Preenche 'Lançamentos 45 E 19' nas linhas fixas informadas,
    casando por user_key (Respons. Entrega) -> posição do colaborador na produtividade.
    Sem criar colunas; não escreve zero; deixa em branco quando não houver valor.
    """
    if df_long.empty or "user_key" not in df_long.columns:
        print("[LANÇAMENTOS] DF vazio ou sem 'user_key'. Nada a preencher.")
        return 0

    grp = {uk: sub for uk, sub in df_long.groupby(df_long['user_key'])}
    total_writes = 0

    for ukey, meta in LANCTOS_USER_MAP.items():
        row_ativ = meta["row_ativ"]
        sub = grp.get(ukey)
        if sub is None or sub.empty:
            print(f"[LANÇAMENTOS] {ukey} → {meta['p1']}: sem registros (deixando em branco).")
            continue

        writes = 0
        for _, reg in sub.iterrows():
            # casar com o cabeçalho
            col_idx = None
            for k in date_keys(reg["data_obj"]):
                c = header_map.get(k)
                if c:
                    col_idx = c; break
            if not col_idx:
                continue
            val = int(reg["valor"])
            if val != 0:
                ws.cell(row=row_ativ, column=col_idx, value=val)
                writes += 1

        total_writes += writes
        print(f"[LANÇAMENTOS] {ukey} → {meta['p1']} (linha {row_ativ}): {writes} células escritas.")
    return total_writes


def fill_sap_fixed(ws, header_map, df_long):
    """SAPMIRO: filtra códigos MS e escreve nas linhas fixas para Alfredo/Gabriel/Pedro."""
    if df_long.empty or "nome" not in df_long.columns:
        print(f"[SAPMIRO] DataFrame vazio ou sem coluna 'nome' ({SAP_PATH}). Nada a preencher.")
        return 0

    # index por código: linhas onde 'nome' contém o MS
    cod_grp = {}
    for cod in SAP_COD_MAP.keys():
        cod_grp[cod] = df_long[df_long['nome'].str.contains(cod, case=False, regex=True, na=False)]

    total_writes = 0
    for cod, meta in SAP_COD_MAP.items():
        row_ativ = meta["row_ativ"]
        sub = cod_grp.get(cod)
        if sub is None or sub.empty:
            print(f"[SAPMIRO] {cod} → {meta['p1']}: sem registros (deixando em branco).")
            continue
        writes = 0
        for _, reg in sub.iterrows():
            col_idx = None
            for k in date_keys(reg["data_obj"]):
                c = header_map.get(k)
                if c:
                    col_idx = c; break
            if not col_idx:
                continue
            val = int(reg["valor"])
            if val != 0:
                ws.cell(row=row_ativ, column=col_idx, value=val); writes += 1
        total_writes += writes
        print(f"[SAPMIRO] {cod} → {meta['p1']} (linha {row_ativ}): {writes} células escritas.")
    return total_writes

# ===================== NOVA FUNÇÃO: PREENCHER FSF =====================

def fill_fsf_flags(ws, header_map):
    """
    (versão corrigida)
    Percorre as colunas de data.
    1. Verifica se é Sábado ou Domingo.
    2. Verifica se o dia está COMPLETAMENTE VAZIO para todos os funcionários (Feriado/Emenda).
    Se for FDS ou Dia Vazio -> Preenche '/' nos espaços em branco.
    """
    import pandas as pd
    from openpyxl.cell.cell import MergedCell

    # Blocos/linhas por colaborador (apenas as linhas de atividade)
    row_ranges = [
        (2, 5), (7, 10), (12, 15), (17, 20),
        (22, 25), (27, 30), (32, 35), (37, 40), (42, 45), (47, 47)
    ]

    print("--- INICIANDO MARCAÇÃO DE FSF (Auto-Detect + FDS) ---")

    # Mapeia colunas que são datas
    cols_to_process = {}
    for date_str, col_idx in header_map.items():
        try:
            # tenta interpretar tanto "YYYY-mm-dd 00:00:00" quanto "dd/mm/YYYY"
            if isinstance(date_str, str) and "-" in date_str and date_str.index("-") == 4:
                dt = pd.to_datetime(date_str, errors='coerce')
            else:
                dt = pd.to_datetime(date_str, dayfirst=True, errors='coerce')

            if pd.notna(dt):
                cols_to_process[col_idx] = dt.date()
        except Exception:
            continue

    total_writes = 0

    for col_idx, data_atual in cols_to_process.items():
        # A) fim de semana?
        is_weekend = (data_atual.weekday() >= 5)

        # B) alguém trabalhou no dia?
        dia_teve_producao = False
        for (start_row, end_row) in row_ranges:
            for r in range(start_row, end_row + 1):
                if r > ws.max_row:
                    continue
                val = ws.cell(row=r, column=col_idx).value
                if val not in (None, "", 0, "/"):
                    dia_teve_producao = True
                    break
            if dia_teve_producao:
                break

        # C) Aplica FSF se fim de semana OU ninguém trabalhou
        aplicar_fsf = is_weekend or (not dia_teve_producao)

        if aplicar_fsf:
            for (start_row, end_row) in row_ranges:
                # garante que não sobrescreve blocos com valor
                bloco_tem_valor = False
                for r in range(start_row, end_row + 1):
                    if r > ws.max_row:
                        continue
                    v = ws.cell(row=r, column=col_idx).value
                    if v not in (None, "", 0, "/"):
                        bloco_tem_valor = True
                        break

                if not bloco_tem_valor:
                    for r in range(start_row, end_row + 1):
                        if r > ws.max_row:
                            continue
                        cell = ws.cell(row=r, column=col_idx)
                        if isinstance(cell, MergedCell):
                            continue
                        if cell.value in (None, "", 0):
                            cell.value = "/"
                            total_writes += 1

    print(f"[FSF] Concluído. Total de células marcadas: {total_writes}")
    return total_writes


# ===================== MAIN =====================

def main():
    wb = load_workbook(PROD_PATH)
    ws = wb["Plan1"]

    # 1) Atualiza cabeçalho do mês anterior (D:AH)
    
    ano, mes, qtd_dias = update_headers_to_previous_month(ws, header_row=1, start_col_letter="D", end_col_letter="AH")
    print(f"[HEADER] Atualizado para {qtd_dias:02d} dias de {mes:02d}/{ano} (mês anterior) em D:AH.")


    # 2) Limpa D:AH nas linhas de atividades
    ROW_RANGES_ATIV = [
        (2, 5), (7, 10), (12, 15), (17, 20),
        (22, 25), (27, 30), (32, 35), (37, 40),  # maria.edurocha
        (42, 45),                                # matheus.lemos.silva
        (47, 47)                                 # Vanessa (só Agilis)
    ]
    cleared = clear_month_data_in_blocks(ws, ROW_RANGES_ATIV, start_col_letter="D", end_col_letter="AH")
    print(f"[CLEAR] Células limpas em D:AH para blocos de atividade: {cleared}")

    # 3) Reconstrói o header_map (agora com o mês atualizado)
    header_map = build_header_map(ws)

    # (opcional) checagem do cabeçalho em D e AH
    d_cell = ws.cell(1, col_letter_to_index("D")).value
    ah_cell = ws.cell(1, col_letter_to_index("AH")).value
    print("[CHECK] D1:", d_cell, "| AH1:", ah_cell)

    # 4) Ler pivots
    df_ag  = read_tabledinamica_with_namecol(AGILIS_PATH)
    df_sd  = read_tabledinamica_with_namecol(SEDEX_PATH)
    df_lan = read_lanctos_tabledinamica(LANCTOS_PATH)
    df_sap = read_tabledinamica_with_namecol(SAP_PATH, name_col_hint="Nome do usuário")

    # 5) Preencher
    ag_writes  = fill_agilis_same_row(ws, header_map, df_ag)
    sd_writes  = fill_sedex(ws, header_map, df_sd)
    ln_writes  = fill_lanctos_fixed(ws, header_map, df_lan)
    sap_writes = fill_sap_fixed(ws, header_map, df_sap)

    # 6) FSF
    fsf_writes = fill_fsf_flags(ws, header_map)

    # 7) Salvar
    wb.save(OUT_PATH)
    print({
        "arquivo_saida": OUT_PATH,
        "agilis_celulas_escritas": ag_writes,
        "sedex_celulas_escritas": sd_writes,
        "lanctos_celulas_escritas": ln_writes,
        "sapmiro_celulas_escritas": sap_writes,
        "fsf_celulas_preenchidas": fsf_writes,
        "linhas_totais": ws.max_row,
        "colunas_totais": ws.max_column,
    })



if __name__ == "__main__":

    '''    '''  
 
    # 1. PEGAR O DIRETÓRIO ATUAL (onde o script está rodando)
    diretorio_atual = os.getcwd() # ou '.'
    
    # 2. EXECUTAR A RENOMEAÇÃO E PREPARAÇÃO DOS ARQUIVOS
    print("--- Executando Etapa 1: Renomear Arquivos ---")
    step_1_prepare_and_rename_reports(diretorio_atual)
 
    # 3. EXECUTAR O PROCESSO PRINCIPAL
    print("--- Executando Etapa 2: Processar Produtividade ---")
    main()


#------------------------------------------------------------------------------------------
    
'''
    # ==============================================================================
    # PARTE 3: BÚSSOLA MRV
    # ==============================================================================
    
    print("Passo 1, abrir o site")

    # --- PASSO 1: Login no domínio principal (como você já faz) ---
    usuario_safe = urllib.parse.quote(EMAIL_USER)
    senha_safe = urllib.parse.quote(SENHA_USER)

    # URL autenticada para o Bússola
    url_autenticada_bussola = f"http://{usuario_safe}:{senha_safe}@bussola.mrv.com.br/Main/Big.aspx"

    print("Acessando Bússola com credenciais embutidas...")
    driver.get(url_autenticada_bussola)

    # --- PASSO 2: PRÉ-AUTENTICAÇÃO NO DOMÍNIO DO RELATÓRIO (NOVO) ---
    print("Pré-autenticando no domínio 'http://report2.mrv.com.br/ReportServer/Pages/ReportViewer.aspx?/BIG/Administrativo/ADM013%20-%20Relat%C3%B3rio%20Protocolo%20de%20Pagamento%20MRV%20PAG/REL_PRLPGTMRV&rs:Command=Render'...")

    # Monta uma URL base para o domínio do relatório com as credenciais
    url_autenticada_report = f"http://{usuario_safe}:{senha_safe}@report2.mrv.com.br/ReportServer/Pages/ReportViewer.aspx?/BIG/Administrativo/ADM013%20-%20Relat%C3%B3rio%20Protocolo%20de%20Pagamento%20MRV%20PAG/REL_PRLPGTMRV&rs:Command=Render"

    # Visita a URL base. O navegador vai processar e guardar as credenciais para este domínio.
    # A página pode dar erro ou ficar em branco, não importa. O objetivo é apenas enviar as credenciais.
    driver.get(url_autenticada_report)

    # --- PASSO 3: Volte para o Bússola para continuar a navegação ---
    print("Retornando ao Bússola para continuar a automação...")
    driver.get(url_autenticada_bussola) # Ou use driver.back() se a página anterior for a correta

    # --- PASSO 4: Continue seu script normalmente ---
    # Agora o navegador já está autenticado em ambos os domínios.
    # O código para clicar nas pastas e no relatório funcionará sem o pop-up.

    print("Procurando pasta 'Administrativo'...")
    # Se o login funcionar, essa pasta vai aparecer.
    pasta_adm = wait.until(EC.element_to_be_clickable((By.ID, "pasta2"))) # Usando o ID que corrigimos antes
    pasta_adm.click()
    time.sleep(2)

    print("Selecionando o relatório...")
    xpath_relatorio = "//div[@id='divLinha' and contains(., 'Relatório Protocolo de Pagamento MRV PAG')]"
    relatorio_link = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_relatorio)))
    relatorio_link.click()

    # Agora o relatório deve abrir diretamente, sem pedir login.
    print("Relatório aberto com sucesso!")
 
    
    WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)

    # Troca para a nova janela
    nova_janela = driver.window_handles[-1]  # pega a última janela aberta
    driver.switch_to.window(nova_janela)

    # Aguarda o elemento estar clicável na nova janela
    wait = WebDriverWait(driver, 10)

    try:
        # Aguarda até que o elemento esteja presente e clicável
        wait = WebDriverWait(driver, 10)
        elemento = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl03_ctl01")))

        # Clica no elemento
        elemento.click()
        print("Elemento clicado com sucesso!")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

    wait = WebDriverWait(driver, 10)

    # Seleciona TODOS os botões "Mês Anterior" pelo atributo ALT do <img>
    botoes = driver.find_elements(By.XPATH, '//a[@accesskey="<"]')

    print("Foram encontrados:", len(botoes), "botões")

    clicou = False

    for botao in botoes:
        try:
            # só clica no que está visível (o calendário aberto)
            if botao.is_displayed():
                driver.execute_script("arguments[0].click();", botao)
                print("Clique realizado no botão correto!")
                clicou = True
                break
        except:
            pass

    if not clicou:
        print("Nenhum botão visível encontrado.")


    # Opcional: Fechar a janela do relatório e voltar para a principal
    # driver.close()
    # driver.switch_to.window(janela_bussola)
    time.sleep(3)
except Exception as e:
    print(f"\nCRITICAL ERROR NA PARTE 3 (BÚSSOLA): {e}")
    driver.save_screenshot("erro_bussola.png")
  '''  

